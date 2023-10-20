from openpyxl import Workbook, load_workbook

wb = Workbook()

ws = wb.active

ws1 = wb.create_sheet('Mysheet')

ws2 = wb.create_sheet('Mysheet2', 0)

ws.title = 'First Title'

ws1['A5'] = 'ABC'

ws2['A4'] = 4

d = ws.cell(row=4, column=2, value=10)

wb.save(f'banana.xlsx')

wb = load_workbook('banana.xlsx')

print(wb.sheetnames)
